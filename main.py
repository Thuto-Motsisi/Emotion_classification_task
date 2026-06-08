import streamlit as st
import pandas as pd
from openpyxl import load_workbook

df = pd.read_excel("emotional_sentences.xlsx", sheet_name='emotion_sentences')
st.title("Emotion Labeling")
# st.write(df)

# Pick 10 unlabelled sentences at the start
if "sentences" not in st.session_state:
    unlabelled = df[df["Emotion Label"].isna()].head(10)
    st.session_state.sentences = unlabelled
    st.session_state.current = 0

sentences = st.session_state.sentences
idx = st.session_state.current
total = len(sentences)
remaining = total - idx - 1

# Get the current sentence
row = sentences.iloc[idx]
sentence = row["Sentence"]
excel_row = row.name + 2  # +2 because Excel rows start at 1 and row 1 is header

st.write(sentence)

col1, col2 = st.columns(2)

with col1:
    left_label = st.radio("What emotion is this?", ["Joy", "Sadness", "Anger", "Fear"], index=None)

with col2:
    right_label = st.radio("", ["Disgust", "Surprise", "Neutral"], index=None)

label = left_label or right_label

confidence = st.slider("Confidence level", min_value=0.0, max_value=1.0, step=0.05)

if label:
    st.write(f"You picked **{label}** with confidence **{confidence}**")


if st.button("Save Label"):
    wb = load_workbook("emotional_sentences.xlsx")
    ws = wb["emotion_sentences"]
    ws["C2"] = label
    ws["D2"] = confidence
    wb.save("emotional_sentences.xlsx")
    st.success("Saved!")

if st.button("Save & Next"):
    wb = load_workbook("emotional_sentences.xlsx")
    ws = wb["emotion_sentences"]
    ws.cell(row=excel_row, column=3, value=label)
    ws.cell(row=excel_row, column=4, value=confidence)
    wb.save("emotional_sentences.xlsx")

    if idx + 1 < total:
        st.session_state.current += 1
        st.rerun()
    else:
        st.success("All 10 sentences labelled!")

st.write(f"{remaining} more sentences to go")


